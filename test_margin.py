#import Openpyxl library
import openpyxl

#load workbook into python
wb = openpyxl.load_workbook('..\\test_margin.xlsx')

#call the active worksheet, which is the first sheet in the workbook
ws = wb.active

#count the number of rows in the worksheet
print('Total number of rows: ', ws.max_row)

#count the number of columns in the worksheet
print('Total number of columns: ', ws.max_column)

#retrieve data from a specific cell, e.g. A3
print('Data in cell A3: ', ws['A3'].value)

#iterate all values in a specific row, e.g row 3
values = [ws.cell(row=2, column=i).value for i in range(1, ws.max_column+1)]
print('Values in row 3: ', values)

#iterate all the values in a specific column, e.g. column B
values = [ws.cell(row=i, column=2).value for i in range(1, ws.max_row+1)]
print('Values in column B: ', values)

#reading data from a range of cells, e.g. B1 to C3
my_list = list()

for value in ws.iter_rows(min_row=2, max_row=3, min_col=1, max_col=3, values_only=True):
    my_list.append(value)

for ele1, ele2, ele3 in my_list:
    print("{:<10}{:<30}{:<20}".format(ele1, ele2, ele3))

#print the name of the active sheet
print('Name of the active sheet: ', ws.title)

#increase font size in cell A1 and bold text
ws['A1'].font = openpyxl.styles.Font(size=14, bold=True)

#change the font size and style for all cells in row 2 to row 3
for cell in ws['2:2']:
    cell.font = openpyxl.styles.Font(size=13, bold=True)

#change the color in A2 to red, bold and font size 13
ws['A2'].font = openpyxl.styles.Font(color='FF0000', bold=True, size=13)

#change the background color in A3 to yellow
ws['A3'].fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

#add a cell border to A4
ws['A4'].border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'), bottom=openpyxl.styles.Side(style='thin'))

#change the scale of the worksheet
ws.page_setup.scale = 65

#change the margin of the worksheet
ws.page_margins.left = 0.15
ws.page_margins.right = 0.15
ws.page_margins.top = 0.6
ws.page_margins.bottom = 0.4

#save the workbook
wb.save('..\\test_margin_saved.xlsx')

